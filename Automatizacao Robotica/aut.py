import pyautogui
import time
import os
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from datetime import datetime
import pygetwindow as gw

pyautogui.PAUSE = 2 


dados = {
    'Tarefa': [],
    'Executou': [],
    'Tempo Execucao': []
}

def abre_programas(nome_programa, tipo):

    dados["Tarefa"].append(f"Abrir {nome_programa}")

    pyautogui.alert("Automação 1")
    inicio = time.time()

    pyautogui.hotkey("win", "s")
    pyautogui.write(nome_programa, interval=0.1)
    pyautogui.press(tipo)
    time.sleep(1)
    

    fim = time.time()
    tempo_execucao = round(fim - inicio)
    dados["Tempo Execucao"].append(tempo_execucao)


    janela = nome_programa
    janelas_encontradas = [win for win in gw.getWindowsWithTitle(janela) if win.visible]

    if janelas_encontradas:
        pyautogui.alert(f"{janela} está aberto!")
        dados["Executou"].append(True)
               
    else:
        pyautogui.alert(f"{janela} não está aberto.")
        dados["Executou"].append(False)


def fazer_pesquisa(navegador, conteudo_pesquisa):

    dados["Tarefa"].append(f"Pesquisando sibre {conteudo_pesquisa}")

    pyautogui.alert("Automação 2")

    inicio = time.time()

    pyautogui.hotkey("win", "s")
    pyautogui.write(navegador)
    pyautogui.press("enter")
    time.sleep(1)


    pyautogui.write(conteudo_pesquisa)
    pyautogui.press("enter")
    time.sleep(1.2)


    pyautogui.click(x=500, y=450)
    

    fim = time.time()
    tempo_execucao = round(fim-inicio)
    
    dados["Tempo Execucao"].append(tempo_execucao)

    janela = navegador
    janelas_encontradas = [win for win in gw.getWindowsWithTitle(janela) if win.visible] #
    
    if janelas_encontradas:
            pyautogui.alert(f'A pesquisa "{conteudo_pesquisa}" foi feita!')
            dados["Executou"].append(True)
            
    else:
        pyautogui.alert(f'A pesquisa "{conteudo_pesquisa}" não foi feita!')
        dados["Executou"].append(False)


def print_tela():

    dados["Tarefa"].append(f"Fazendo captura de tela")

    pyautogui.alert("Automação 3")

    inicio = time.time()

    pyautogui.alert("Clique para selecionar o inicio da área da captura")
    time.sleep(5)
    x1, y1 = pyautogui.position()

    pyautogui.alert("Clique para selecionar o fim da área de captura")
    time.sleep(5)
    x2, y2 = pyautogui.position()

    largura = abs(x2 - x1)
    altura = abs(y2 - y1)
    topo = min(y1, y2)
    esquerda = min(x1, x2)

    captura = pyautogui.screenshot(region=(esquerda, topo, largura, altura))

    nome_captura = datetime.now().strftime("captura_%Y%m%d_%H%M%S.png")
    captura.save(nome_captura)
    
    try:
        os.startfile(nome_captura)
        dados["Executou"].append(True)

    except:
        pyautogui.alert("Erro ao realizar a captura")
        dados["Executou"].append(False)

    fim = time.time()
    tempo_execucao = round(fim-inicio)
    dados["Tempo Execucao"].append(tempo_execucao)

    time.sleep(2)


def olha_galeria(nome_programa):

    dados["Tarefa"].append(f"Abrindo {nome_programa}")
    pyautogui.alert("Automação 4")
    inicio = time.time()

    pyautogui.hotkey("win", "s")
    pyautogui.write(nome_programa, interval=0.1)
    pyautogui.press("enter")
    time.sleep(2)

    pyautogui.click(x=500, y=450)
    time.sleep(2)

    pyautogui.press("f11")
    time.sleep(1)

    quantidade = pyautogui.prompt("Quantas fotos deseja ver?")
    try:
        quantidade = int(quantidade)
    except ValueError:
        pyautogui.alert("Valor inválido! Encerrando tarefa.")
        dados["Executou"].append(False)
        dados["Tempo Execucao"].append(0)
        return

    if quantidade <= 0:
        pyautogui.alert("Fechando o programa \nMotivo: Número menor ou igual a zero")
        dados["Executou"].append(False)
    else:
        for i in range(quantidade):
            pyautogui.press("right")
            time.sleep(1.3)
        dados["Executou"].append(True)

    fim = time.time()
    tempo_execucao = round(fim - inicio)
    dados["Tempo Execucao"].append(tempo_execucao)
   

df = pd.read_csv("tarefasAutomatizadas.csv", encoding='utf-8')

for index, linha in df.iterrows():
    tarefa = linha["Tarefa"].strip().lower()
    tipo = str(linha["Tipo"]).strip().lower()
    dado = str(linha["Dado"]).strip()

    if "abrir" in tarefa:
        abre_programas(dado, tipo)
        time.sleep(5)

    elif "capturar" in tarefa:
        print_tela()
        time.sleep(5)

    elif "pesquisar" in tarefa:
        fazer_pesquisa("Google", dado)
        time.sleep(5)
    
    elif "mostrar galeria" in tarefa:
        olha_galeria(dado)
        time.sleep(5)

    else:
        pyautogui.alert(f"Tarefa desconhecida: {tarefa}")



pyautogui.alert("Automações finalizadas, obrigado por rodar nosso programa")
print(dados)

arquivo_excel = "tarefas-realizadas.xlsx"
df_tarefas = pd.DataFrame(dados)


if os.path.exists(arquivo_excel):
    wb = load_workbook(arquivo_excel)
    ws = wb.active
    for row in dataframe_to_rows(df_tarefas, index=False, header=False):
        ws.append(row)
else:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tarefas"

    for row in dataframe_to_rows(df_tarefas, index=False, header=True):
        ws.append(row)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

wb.save(arquivo_excel)

